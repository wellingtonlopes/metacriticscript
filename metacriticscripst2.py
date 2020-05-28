import requests
import os.path
from config import api_headers
from openpyxl import Workbook, load_workbook


if os.path.isfile('gamescoresap.xlsx'):
    scorebook = load_workbook(filename='gamescoresap.xlsx')
else:
    scorebook = Workbook()
    scoresheet_pc = scorebook.active
    scoresheet_pc.title = "Game Pass for PC Metascores"
    scoresheet_pc.append(["Game Title", "Metascore"])
    scoresheet_xbox = scorebook.create_sheet("Game Pass for Xbox Metascores")
    scoresheet_xbox.append(["Game Title", "Metascore"])
    scoresheet_switch = scorebook.create_sheet("Switch Metascores")
    scoresheet_switch.append(["Game Title", "Metascore"])
    scoresheet_ps4 = scorebook.create_sheet("PS4 Metascores")
    scoresheet_ps4.append(["Game Title", "Metascore"])


def get_games():
    game_source = load_workbook(filename="game_source.xlsx")
    gs_sheet = game_source.active
    game_titles = gs_sheet['A']
    game_platform = gs_sheet['B']
    game_url_platform = []
    for (game, platform) in zip(game_titles[1:], game_platform[1:]):
        if platform.value == "Both":
            url_game = "%20".join(game.value.split())
            game_url_platform.append((url_game, "pc"))
            game_url_platform.append((url_game, "xbox-one"))
        elif platform.value == "Xbox 360":
            url_game = "%20".join(game.value.split())
            game_url_platform.append((url_game, "xbox-360"))
        elif platform.value == "Xbox One":
            url_game = "%20".join(game.value.split())
            game_url_platform.append((url_game, "xbox-one"))
        elif platform.value == "PC":
            url_game = "%20".join(game.value.split())
            game_url_platform.append((url_game, "pc"))
        elif platform.value == "Switch":
            url_game = "%20".join(game.value.split())
            game_url_platform.append((url_game, "switch"))
        elif platform.value == "PS4":
            url_game = "%20".join(game.value.split())
            game_url_platform.append((url_game, "playstation-4"))
            
    return game_url_platform


def metascore():
    games_platform = get_games()
    api_url = "https://chicken-coop.p.rapidapi.com/games/"
    platforms = [platform for title, platform in games_platform]
    urls = [ api_url + title for title, platform in games_platform]
    headers = api_headers
    keys = ['title', 'score']

    for (url, platform) in zip(urls, platforms):
            response = requests.request("GET", url, headers=headers, params=platform)
            game_info = response.json().get('result')
            game_score = [game_info.get(key) for key in keys] if game_info != 'No result' else [url, "No Data"]
            game_score.append(platform)
            if platform == "pc":
                scoresheet_pc.append(game_score)
            elif platform == "xbox-one" or platform == "xbox-360":
                scoresheet_xbox.append(game_score)
            elif platform == "switch":
                scoresheet_switch.append(game_score)
            else:
                scoresheet_ps4.append(game_score)
            print(game_score)
        
    scorebook.save("gamescoresap.xlsx")


metascore()