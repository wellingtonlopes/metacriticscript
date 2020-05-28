import requests
import os.path
from config import api_headers
from openpyxl import Workbook, load_workbook

# text = open("gameslist.txt", "r")   


if os.path.isfile('gamescores.xlsx'):
    scorebook = load_workbook(filename='gamescores.xlsx')
else:
    scorebook = Workbook()
    scoresheet_pc = scorebook.active
    scoresheet_pc.title = "Game Pass for PC Metascores"
    scoresheet_pc.append(["Game Title", "Metascore"])
    scoresheet_xbox = scorebook.create_sheet("Game Pass for Xbox Metascores")
    scoresheet_xbox.append(["Game Title", "Metascore"])


def get_games_pc():
    game_source = load_workbook(filename="game_source.xlsx")
    gs_sheet = game_source.active
    game_list = gs_sheet['A']
    pc_games_in_url = []
    count = 1
    for game in game_list[1:]:
        count += 1
        game_platform = gs_sheet.cell(row=count, column=2)
        if game_platform.value != "Xbox One":
            url_game = "%20".join(game.value.split())
            pc_games_in_url.append(url_game)

    return pc_games_in_url


def get_games_xbox():
    game_source = load_workbook(filename="game_source.xlsx")
    gs_sheet = game_source.active
    game_list = gs_sheet['A']
    xbox_games_in_url = []
    count = 1
    for game in game_list[1:]:
        count += 1
        game_platform = gs_sheet.cell(row=count, column=2)
        if game_platform.value != "PC":
            url_game = "%20".join(game.value.split())
            xbox_games_in_url.append(url_game)

    return xbox_games_in_url


"""def convert_to_games(text):
    games = text.read().splitlines()
    games_in_url = []
    for game in games:
        if game.endswith("A)"):
            game = game[:-5]
        elif game.endswith(")"):
            game = game[:-6]
        game = "%20".join(game.split())
        games_in_url.append(game)
    return games_in_url
"""


def metascore_pc():
    games = get_games_pc()
    api_url = "https://chicken-coop.p.rapidapi.com/games/"
    api_querystring_pc = {"platform":"pc"}
    urls = [ api_url + game for game in games]
    headers = api_headers
    keys = ['title', 'score']

    for url in urls:
            response = requests.request("GET", url, headers=headers, params=api_querystring_pc)
            game_info = response.json().get('result')
            game_score = [game_info.get(key) for key in keys] if game_info != 'No result' else ["No Data", "No Data"]
            scoresheet_pc.append(game_score)
            print(game_score)
        
    scorebook.save("gamescores.xlsx")


def metascore_xbox():
    games = get_games_xbox()
    api_url = "https://chicken-coop.p.rapidapi.com/games/"
    api_querystring_xbox = {"platform":"xbox-one"}
    urls = [ api_url + game for game in games]
    headers = api_headers
    keys = ['title', 'score']

    for url in urls:
            response = requests.request("GET", url, headers=headers, params=api_querystring_xbox)
            game_info = response.json().get('result')
            game_score = [game_info.get(key) for key in keys] if game_info != 'No result' else ["No Data", "No Data"]
            scoresheet_xbox.append(game_score)
            print(game_score)
        
    scorebook.save("gamescores.xlsx")


metascore_xbox()
metascore_pc()